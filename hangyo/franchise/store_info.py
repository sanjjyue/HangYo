import os
from openpyxl import load_workbook
# from franchise.models import stores

path = "./files"
file_list = os.listdir(path)

for file_name_raw in file_list:
    file_name = "./files/" + file_name_raw
    wb = load_workbook(filename=file_name, data_only=True)
    ws = wb.active
    for i in range(2,10):
        result = []
        # cell_num1 = "A" + str(i)
        cell_num2 = "B" + str(i)
        cell_num3 = "C" + str(i)
        cell_num4 = "D" + str(i)
        cell_num5 = "E" + str(i)
        result.append(ws[cell_num2].value)
        result.append(ws[cell_num3].value)
        result.append(ws[cell_num4].value)
        result.append(ws[cell_num5].value)

#         print(result)
        # stores.objects.create(store_name = ws[cell_num2].value, phone_num = ws[cell_num3].value, location = ws[cell_num4].value, store_type = ws[cell_num5])
       
        print(result)
